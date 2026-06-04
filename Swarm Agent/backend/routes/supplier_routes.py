"""
routes/supplier_routes.py — /api/suppliers/* router.

Each handler filters by ``current_user["client_id"]`` from the JWT. Shared
state (``clients_db``, ``users_db``, ``SEED_CLIENT_IDS``) and helpers
(``_resolve_suppliers``, ``_supplier_limit``, ``_normalize_zone``,
``_now_utc``, ``_save_local_state``, ``_mark_onboarding_step``,
``_create_notification``, ``_ensure_client_defaults``, ``_client_user``,
``_make_suppliers_workbook``, ``VALID_ZONES``) remain in main.py and are
imported lazily inside each handler to avoid a circular import.
"""

import csv
import hashlib
import io
import logging
import random
import secrets
from datetime import datetime, timedelta, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse

import agents
import auth
import email_service
import esg_signals
import financial_signals
import storage
from models import SupplierInput, BulkDeleteRequest

# Walk up from routes/supplier_routes.py until we find a directory containing "dataset/"
# routes/ -> backend/ -> Swarm Agent/ -> repo-root/   (3 levels up on Railway: /app/)
def _find_dataset_dir() -> Path:
    candidate = Path(__file__).resolve()
    for _ in range(6):
        candidate = candidate.parent
        if (candidate / "dataset").is_dir():
            return candidate / "dataset"
    # Last resort: same directory as this file
    return Path(__file__).resolve().parent / "dataset"

_DATASET_DIR = _find_dataset_dir()

SAMPLE_DATASETS = [
    {"id": "01", "filename": "01_Automotive_Global_25_suppliers.xlsx",       "industry": "Automotive",       "geography": "Global", "supplier_count": 25, "description": "Tier-1 and Tier-2 automotive parts suppliers across North America, Europe, and Asia."},
    {"id": "02", "filename": "02_Electronics_Global_35_suppliers.xlsx",      "industry": "Electronics",      "geography": "Global", "supplier_count": 35, "description": "Semiconductor, PCB, and consumer electronics component suppliers worldwide."},
    {"id": "03", "filename": "03_Pharma_India_18_suppliers.xlsx",            "industry": "Pharmaceutical",   "geography": "India",  "supplier_count": 18, "description": "Active Pharmaceutical Ingredient (API) and excipient suppliers across India."},
    {"id": "04", "filename": "04_FMCG_Mixed_40_suppliers.xlsx",              "industry": "FMCG",             "geography": "Mixed",  "supplier_count": 40, "description": "Fast-moving consumer goods suppliers spanning packaging, ingredients, and logistics."},
    {"id": "05", "filename": "05_Aerospace_Global_22_suppliers.xlsx",        "industry": "Aerospace",        "geography": "Global", "supplier_count": 22, "description": "Aerospace-grade components, avionics, and MRO suppliers with defence-grade reliability data."},
    {"id": "06", "filename": "06_Renewable_Energy_Global_28_suppliers.xlsx", "industry": "Renewable Energy", "geography": "Global", "supplier_count": 28, "description": "Solar panels, wind turbine components, and energy storage material suppliers."},
    {"id": "07", "filename": "07_Food_Beverage_India_32_suppliers.xlsx",     "industry": "Food & Beverage",  "geography": "India",  "supplier_count": 32, "description": "Agricultural commodities, processing equipment, and cold-chain suppliers across India."},
    {"id": "08", "filename": "08_Chemicals_Mixed_20_suppliers.xlsx",         "industry": "Chemicals",        "geography": "Mixed",  "supplier_count": 20, "description": "Specialty chemicals, solvents, and industrial compounds suppliers across key regions."},
    {"id": "09", "filename": "09_Logistics_3PL_Global_15_suppliers.xlsx",    "industry": "Logistics / 3PL",  "geography": "Global", "supplier_count": 15, "description": "Third-party logistics, freight, and last-mile delivery network providers."},
    {"id": "10", "filename": "10_Medical_Devices_Global_38_suppliers.xlsx",  "industry": "Medical Devices",  "geography": "Global", "supplier_count": 38, "description": "Medical device components, sterilisation materials, and regulatory-compliant OEM suppliers."},
]

SAMPLE_DATASET_FILENAMES = {d["filename"] for d in SAMPLE_DATASETS}

logger = logging.getLogger("disruptiq.routes.suppliers")

supplier_router = APIRouter(prefix="/api/suppliers", tags=["suppliers"])


@supplier_router.get("/sample-datasets/debug-path")
async def debug_dataset_path():
    """Debug: show resolved dataset directory and which files exist. Public."""
    import os
    file_here = str(Path(__file__).resolve())
    exists = _DATASET_DIR.exists()
    files = sorted(os.listdir(_DATASET_DIR)) if exists else []
    return {
        "this_file": file_here,
        "dataset_dir": str(_DATASET_DIR),
        "dataset_dir_exists": exists,
        "files_found": files,
    }


@supplier_router.get("/sample-datasets")
async def list_sample_datasets():
    """Return the catalogue of DisruptIQ sample datasets. Public — no auth required."""
    return {"datasets": SAMPLE_DATASETS}


@supplier_router.get("/sample-datasets/{filename}")
async def download_sample_dataset(filename: str, request: Request):
    """Stream a sample dataset .xlsx file. Public — no auth required for download."""
    # Guard against path traversal
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    matched = next((d for d in SAMPLE_DATASETS if d["filename"] == filename), None)
    if not matched:
        raise HTTPException(status_code=404, detail="Sample dataset not found")
    file_path = _DATASET_DIR / filename
    if not file_path.exists():
        logger.warning("Sample dataset file missing on server: %s (looked in %s)", filename, _DATASET_DIR)
        raise HTTPException(status_code=404, detail="Dataset file not available on this server")
    # Best-effort audit log — only when caller is authenticated
    try:
        user = await auth.get_optional_user(request)
        if user:
            storage.write_audit(
                event_id=f"sample_dataset_{user.get('client_id')}",
                agent="OnboardingSystem",
                action="sample_dataset_downloaded",
                input_summary=filename,
                output_summary=f"industry={matched['industry']} count={matched['supplier_count']}",
                client_id=user.get("client_id"),
            )
    except Exception:
        pass
    return StreamingResponse(
        open(file_path, "rb"),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@supplier_router.get("/health")
async def suppliers_health(current_user: dict = Depends(auth.require_auth)):
    """Light health check - confirms the suppliers router is mounted."""
    return {
        "status": "ok",
        "router": "supplier_routes",
        "client_id": current_user.get("client_id"),
    }


@supplier_router.get("")
async def suppliers_route(current_user: dict = Depends(auth.require_auth)):
    """Get current user's suppliers."""
    from main import _resolve_suppliers
    client_id = current_user.get("client_id")
    if not client_id:
        raise HTTPException(status_code=400, detail="User has no associated client")
    return {"client_id": client_id, "suppliers": _resolve_suppliers(client_id)}


@supplier_router.get("/template")
async def download_supplier_template():
    """Generate and return an Excel template for supplier onboarding. No auth required."""
    from main import VALID_ZONES
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    headers = [
        "Supplier Name*",
        "Zone*",
        "Categories* (comma-separated)",
        "Buffer Stock Days",
        "Sites",
        "Reliability (%)",
        "Proximity Score (1-10)",
    ]
    header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    col_widths = [28, 20, 36, 20, 10, 18, 22]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32

    example_rows = [
        ["[Example] Supplier Alpha", "Mumbai", "raw materials, components", 14, 3, 92, 8],
        ["[Example] Supplier Beta", "Pune", "logistics, packaging", 7, 2, 85, 6],
        ["[Example] Supplier Gamma", "Bengaluru", "specialty chemicals", 18, 4, 94, 3],
    ]
    row_fill = PatternFill(start_color="0D1123", end_color="0D1123", fill_type="solid")
    row_font = Font(color="E2E8F0", size=10)
    row_alignment = Alignment(vertical="center")

    for row_idx, row_data in enumerate(example_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = row_font
            cell.fill = row_fill
            cell.alignment = row_alignment

    notes_row_idx = len(example_rows) + 2
    notes_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    notes_font = Font(color="93C5FD", size=9, italic=True)
    notes_text = (
        f"NOTES: Valid zones: {', '.join(VALID_ZONES)}. "
        "Reliability: 0-100 (integer). Proximity Score: 1-10 (integer). "
        "Buffer Stock Days: integer (e.g. 7). Sites: integer (e.g. 2). "
        "Categories: comma-separated text. Max 50 suppliers. * = required."
    )
    notes_cell = ws.cell(row=notes_row_idx, column=1, value=notes_text)
    notes_cell.font = notes_font
    notes_cell.fill = notes_fill
    notes_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(
        start_row=notes_row_idx, start_column=1,
        end_row=notes_row_idx, end_column=len(headers)
    )
    ws.row_dimensions[notes_row_idx].height = 48
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=disruptiq_supplier_template.xlsx"},
    )


@supplier_router.post("/upload-excel")
async def upload_supplier_excel(
    request: Request,
    file: UploadFile = File(...),
    current_user: dict = Depends(auth.require_auth),
):
    """Parse an uploaded Excel file and store suppliers for the authenticated client."""
    from main import (clients_db, SEED_CLIENT_IDS, _supplier_limit, _normalize_zone,
                      _now_utc, _save_local_state, _mark_onboarding_step,
                      _create_notification)
    MAX_FILE_SIZE = 10 * 1024 * 1024
    client_id = current_user.get("client_id")
    if client_id in SEED_CLIENT_IDS:
        raise HTTPException(status_code=403, detail="Seed/demo accounts cannot upload suppliers.")
    client = clients_db.get(client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx and .xls files are accepted")

    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail=f"File size exceeds {MAX_FILE_SIZE // (1024*1024)} MB limit")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        logger.warning("Excel parse error for client %s: %s", client_id, e)
        raise HTTPException(status_code=400, detail="Could not parse Excel file. Ensure it is a valid .xlsx/.xls file.")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel file has no data rows")

    header_row_idx = None
    for scan_idx, row in enumerate(rows[:10]):
        normalized = [
            str(h).lower().strip().split("(")[0].rstrip("* ").strip() if h else ""
            for h in row
        ]
        if any("supplier name" in h or h == "name" for h in normalized) and \
           any("zone" in h for h in normalized):
            header_row_idx = scan_idx
            break

    if header_row_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Required columns 'Supplier Name' and 'Zone' not found in the first 10 rows",
        )

    raw_headers = [
        str(h).lower().strip().split("(")[0].rstrip("* ").strip() if h else ""
        for h in rows[header_row_idx]
    ]
    rows = rows[header_row_idx + 1:]

    col_map = {}
    for idx, h in enumerate(raw_headers):
        if "supplier name" in h or h == "name":
            col_map["name"] = idx
        elif "zone" in h:
            col_map["zone"] = idx
        elif "categor" in h:
            col_map["categories"] = idx
        elif "buffer" in h:
            col_map["buffer_stock_days"] = idx
        elif "sites" in h:
            col_map["sites"] = idx
        elif "reliability" in h:
            col_map["reliability"] = idx
        elif "proximity" in h:
            col_map["proximity_score"] = idx

    if "name" not in col_map or "zone" not in col_map:
        raise HTTPException(
            status_code=400,
            detail="Required columns 'Supplier Name' and 'Zone' not found",
        )

    existing_suppliers = client.get("suppliers", [])
    existing_count = len(existing_suppliers)
    slots_remaining = _supplier_limit(client_id) - existing_count

    if slots_remaining <= 0:
        return {
            "success": True,
            "suppliers_added": 0,
            "total_suppliers": existing_count,
            "skipped": -1,
            "limit_reached": True,
            "errors": [],
            "message": (
                f"You already have {existing_count} suppliers - the free-plan limit of 30 is reached. "
                "No new suppliers were imported. To add more than 30 suppliers, upgrade to "
                "DisruptIQ Premium or contact kcsbadp@gmail.com."
            ),
        }

    new_suppliers = []
    errors = []
    skipped = 0

    def _safe_int(val, default, lo=None, hi=None):
        try:
            v = int(float(val)) if val else default
            if lo is not None:
                v = max(v, lo)
            if hi is not None:
                v = min(v, hi)
            return v
        except (TypeError, ValueError):
            return default

    for row_idx, row in enumerate(rows, header_row_idx + 2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        raw_name = str(row[col_map["name"]] or "").strip()
        raw_zone = str(row[col_map["zone"]] or "").strip()

        if raw_name.upper().startswith("NOTES") or raw_zone.upper().startswith("NOTES"):
            continue
        if not raw_name:
            errors.append(f"Row {row_idx}: Missing supplier name - skipped")
            continue
        if not raw_zone:
            errors.append(f"Row {row_idx}: Missing zone for '{raw_name}' - skipped")
            continue

        zone = _normalize_zone(raw_zone) or raw_zone.strip()
        raw_cats = row[col_map["categories"]] if "categories" in col_map else None
        categories = (
            [c.strip() for c in str(raw_cats).split(",") if c.strip()]
            if raw_cats
            else ["general"]
        )

        if len(new_suppliers) >= slots_remaining:
            skipped += 1
            continue

        new_suppliers.append({
            "id": f"SUP-{existing_count + len(new_suppliers) + 1:03d}",
            "name": raw_name,
            "zone": zone,
            "buffer_stock_days": _safe_int(
                row[col_map["buffer_stock_days"]] if "buffer_stock_days" in col_map else None, 7, lo=0),
            "sites": _safe_int(
                row[col_map["sites"]] if "sites" in col_map else None, 1, lo=1),
            "reliability": _safe_int(
                row[col_map["reliability"]] if "reliability" in col_map else None, 80, lo=0, hi=100),
            "categories": categories,
            "proximity_score": _safe_int(
                row[col_map["proximity_score"]] if "proximity_score" in col_map else None, 5, lo=1, hi=10),
        })

    if not new_suppliers and skipped == 0:
        raise HTTPException(status_code=400, detail="No valid supplier rows found")

    limit_reached = skipped > 0
    suppliers = existing_suppliers + new_suppliers
    client["suppliers"] = suppliers
    client["supplier_count"] = len(suppliers)
    client["updated_at"] = _now_utc()
    _mark_onboarding_step(client_id, "suppliers_imported", True)
    if file.filename in SAMPLE_DATASET_FILENAMES:
        client["used_sample_dataset"] = True
    _save_local_state()

    is_sample = file.filename in SAMPLE_DATASET_FILENAMES
    storage.write_audit(
        event_id=f"registration_{client_id}",
        agent="OnboardingSystem",
        action="supplier_data_uploaded",
        input_summary=f"added={len(new_suppliers)} total={len(suppliers)} | file={file.filename} | sample={is_sample}",
        output_summary=f"Successfully imported {len(new_suppliers)} new suppliers",
        client_id=client_id,
    )
    _create_notification(client_id, "import_success", "Supplier import complete",
                         f"{len(new_suppliers)} new suppliers added. Total: {len(suppliers)}/30.", "/map")

    user_email = current_user.get("email", "")
    company_name = client.get("company_name", "")
    email_sent = email_service.send_supplier_import_confirmation(
        email=user_email, company_name=company_name, suppliers=new_suppliers, warnings=errors)

    if limit_reached:
        msg = (
            f"Added {len(new_suppliers)} supplier(s) up to the free-plan limit of 30. "
            f"{skipped} additional supplier(s) in your file were not imported. "
            "To add more than 30 suppliers, upgrade to DisruptIQ Premium or contact kcsbadp@gmail.com."
        )
    else:
        msg = f"Added {len(new_suppliers)} suppliers. You now have {len(suppliers)} of 30."
    return {
        "success": True,
        "suppliers_added": len(new_suppliers),
        "total_suppliers": len(suppliers),
        "skipped": skipped,
        "limit_reached": limit_reached,
        "errors": errors,
        "email_sent": email_sent,
        "message": msg,
    }


@supplier_router.post("/upload-csv")
async def upload_supplier_csv(
    request: Request,
    file: UploadFile = File(...),
    current_user: dict = Depends(auth.require_auth),
):
    """Parse an uploaded CSV file and store suppliers for the authenticated client."""
    from main import (clients_db, SEED_CLIENT_IDS, _supplier_limit, _normalize_zone,
                      _now_utc, _save_local_state, _mark_onboarding_step,
                      _create_notification)
    MAX_FILE_SIZE = 10 * 1024 * 1024
    client_id = current_user.get("client_id")
    if client_id in SEED_CLIENT_IDS:
        raise HTTPException(status_code=403, detail="Seed/demo accounts cannot upload suppliers.")
    client = clients_db.get(client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only .csv files are accepted by this endpoint")

    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail=f"File size exceeds {MAX_FILE_SIZE // (1024 * 1024)} MB limit")

    try:
        text = contents.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = contents.decode("latin-1")
        except Exception:
            raise HTTPException(status_code=400, detail="Could not decode CSV file. Please save as UTF-8.")

    reader = csv.DictReader(io.StringIO(text))
    raw_headers = [h.lower().strip().split("(")[0].rstrip("* ").strip() for h in (reader.fieldnames or [])]
    col_name = next((h for h in raw_headers if "supplier name" in h or h == "name"), None)
    col_zone = next((h for h in raw_headers if "zone" in h), None)
    if not col_name or not col_zone:
        raise HTTPException(
            status_code=400,
            detail="Required columns 'Supplier Name' and 'Zone' not found in CSV header",
        )

    def _safe_int_csv(val, default, lo=None, hi=None):
        try:
            v = int(float(str(val).strip())) if val and str(val).strip() else default
            if lo is not None:
                v = max(v, lo)
            if hi is not None:
                v = min(v, hi)
            return v
        except (TypeError, ValueError):
            return default

    existing_suppliers = client.get("suppliers", [])
    existing_count = len(existing_suppliers)
    slots_remaining = _supplier_limit(client_id) - existing_count

    if slots_remaining <= 0:
        return {
            "success": True,
            "suppliers_added": 0,
            "total_suppliers": existing_count,
            "skipped": -1,
            "limit_reached": True,
            "errors": [],
            "message": (
                f"You already have {existing_count} suppliers - the free-plan limit of 30 is reached. "
                "No new suppliers were imported. To add more than 30 suppliers, upgrade to "
                "DisruptIQ Premium or contact kcsbadp@gmail.com."
            ),
        }

    col_cats = next((h for h in raw_headers if "categor" in h), None)
    col_buf = next((h for h in raw_headers if "buffer" in h), None)
    col_sites = next((h for h in raw_headers if h == "sites"), None)
    col_rel = next((h for h in raw_headers if "reliability" in h), None)
    col_prox = next((h for h in raw_headers if "proximity" in h), None)

    new_suppliers = []
    errors = []
    skipped = 0

    for row_idx, raw_row in enumerate(reader, 2):
        row = {k.lower().strip().split("(")[0].rstrip("* ").strip(): v for k, v in raw_row.items()}

        if all(not str(v).strip() for v in row.values()):
            continue

        raw_name = str(row.get(col_name, "") or "").strip()
        raw_zone = str(row.get(col_zone, "") or "").strip()

        if raw_name.upper().startswith("NOTES") or raw_zone.upper().startswith("NOTES"):
            continue
        if not raw_name:
            errors.append(f"Row {row_idx}: Missing supplier name - skipped")
            continue
        if not raw_zone:
            errors.append(f"Row {row_idx}: Missing zone for '{raw_name}' - skipped")
            continue

        zone = _normalize_zone(raw_zone) or raw_zone.strip()
        raw_cats = row.get(col_cats, "") if col_cats else ""
        categories = [c.strip() for c in str(raw_cats).split(",") if c.strip()] or ["general"]

        if len(new_suppliers) >= slots_remaining:
            skipped += 1
            continue

        new_suppliers.append({
            "id": f"SUP-{existing_count + len(new_suppliers) + 1:03d}",
            "name": raw_name,
            "zone": zone,
            "buffer_stock_days": _safe_int_csv(row.get(col_buf, ""), 7, lo=0),
            "sites": _safe_int_csv(row.get(col_sites, ""), 1, lo=1),
            "reliability": _safe_int_csv(row.get(col_rel, ""), 80, lo=0, hi=100),
            "categories": categories,
            "proximity_score": _safe_int_csv(row.get(col_prox, ""), 5, lo=1, hi=10),
        })

    if not new_suppliers and skipped == 0:
        raise HTTPException(status_code=400, detail="No valid supplier rows found in CSV")

    limit_reached = skipped > 0
    suppliers = existing_suppliers + new_suppliers
    client["suppliers"] = suppliers
    client["supplier_count"] = len(suppliers)
    client["updated_at"] = _now_utc()
    _mark_onboarding_step(client_id, "suppliers_imported", True)
    if file.filename in SAMPLE_DATASET_FILENAMES:
        client["used_sample_dataset"] = True
    _save_local_state()

    is_sample = file.filename in SAMPLE_DATASET_FILENAMES
    storage.write_audit(
        event_id=f"registration_{client_id}",
        agent="OnboardingSystem",
        action="supplier_csv_file_uploaded",
        input_summary=f"added={len(new_suppliers)} total={len(suppliers)} | file={file.filename} | sample={is_sample}",
        output_summary=f"Successfully imported {len(new_suppliers)} new suppliers from CSV",
        client_id=client_id,
    )
    _create_notification(client_id, "import_success", "Supplier import complete",
                         f"{len(new_suppliers)} new suppliers added from CSV. Total: {len(suppliers)}/30.", "/map")

    user_email = current_user.get("email", "")
    company_name = client.get("company_name", "")
    email_service.send_supplier_import_confirmation(
        email=user_email, company_name=company_name, suppliers=new_suppliers, warnings=errors)

    if limit_reached:
        msg = (
            f"Added {len(new_suppliers)} supplier(s) up to the free-plan limit of 30. "
            f"{skipped} additional supplier(s) in your file were not imported. "
            "To add more than 30 suppliers, upgrade to DisruptIQ Premium or contact kcsbadp@gmail.com."
        )
    else:
        msg = f"Added {len(new_suppliers)} suppliers. You now have {len(suppliers)} of 30."
    return {
        "success": True,
        "suppliers_added": len(new_suppliers),
        "total_suppliers": len(suppliers),
        "skipped": skipped,
        "limit_reached": limit_reached,
        "errors": errors,
        "message": msg,
    }


@supplier_router.post("/add-single")
async def add_single_supplier(req: SupplierInput, current_user: dict = Depends(auth.require_auth)):
    from main import (SEED_CLIENT_IDS, _client_user, _supplier_limit,
                      _normalize_zone, _now_utc, _save_local_state)
    if current_user.get("client_id") in SEED_CLIENT_IDS:
        raise HTTPException(status_code=403, detail="Seed/demo accounts cannot add suppliers.")
    user, client = _client_user(current_user)
    suppliers = client.setdefault("suppliers", [])
    _limit = _supplier_limit(user["client_id"])
    if len(suppliers) >= _limit:
        raise HTTPException(
            status_code=422,
            detail=(
                f"You have reached the {_limit}-supplier free-plan limit. "
                "To add more, request Premium access in your account settings."
            ),
        )
    zone = _normalize_zone(req.zone) or req.zone.strip()
    supplier = {
        "id": f"SUP-{secrets.token_hex(6).upper()}",
        "name": req.name.strip(),
        "zone": zone,
        "categories": req.categories,
        "buffer_stock_days": max(0, req.buffer_stock_days),
        "sites": max(1, req.sites),
        "reliability": max(0, min(100, req.reliability)),
        "proximity_score": max(1, min(10, req.proximity_score)),
    }
    suppliers.append(supplier)
    client["supplier_count"] = len(suppliers)
    client["updated_at"] = _now_utc()
    _save_local_state()
    storage.write_audit(
        event_id=f"suppliers_{user['client_id']}",
        agent="SupplierManager",
        action="supplier_added",
        input_summary=supplier["name"],
        output_summary=supplier["id"],
        client_id=user["client_id"],
    )
    return {"success": True, "supplier": supplier, "total_count": len(suppliers)}


@supplier_router.put("/{supplier_id}")
async def update_single_supplier(supplier_id: str, req: SupplierInput,
                                 current_user: dict = Depends(auth.require_auth)):
    from main import _client_user, _normalize_zone, _now_utc
    user, client = _client_user(current_user)
    suppliers = client.setdefault("suppliers", [])
    supplier = next((item for item in suppliers if item.get("id") == supplier_id), None)
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    supplier.update({
        "name": req.name.strip(),
        "zone": _normalize_zone(req.zone) or "Bengaluru",
        "categories": req.categories,
        "buffer_stock_days": max(0, req.buffer_stock_days),
        "sites": max(1, req.sites),
        "reliability": max(0, min(100, req.reliability)),
        "proximity_score": max(1, min(10, req.proximity_score)),
    })
    client["updated_at"] = _now_utc()
    storage.write_audit(
        event_id=f"suppliers_{user['client_id']}",
        agent="SupplierManager",
        action="supplier_updated",
        input_summary=supplier_id,
        output_summary=supplier["name"],
        client_id=user["client_id"],
    )
    return {"success": True, "updated_supplier": supplier}


@supplier_router.delete("/{supplier_id}")
async def delete_single_supplier(supplier_id: str, current_user: dict = Depends(auth.require_auth)):
    from main import _client_user, _now_utc
    user, client = _client_user(current_user)
    suppliers = client.setdefault("suppliers", [])
    remaining = [item for item in suppliers if item.get("id") != supplier_id]
    if len(remaining) == len(suppliers):
        raise HTTPException(status_code=404, detail="Supplier not found")
    client["suppliers"] = remaining
    client["supplier_count"] = len(remaining)
    client["updated_at"] = _now_utc()
    storage.write_audit(
        event_id=f"suppliers_{user['client_id']}",
        agent="SupplierManager",
        action="supplier_deleted",
        input_summary=supplier_id,
        output_summary=f"remaining={len(remaining)}",
        client_id=user["client_id"],
    )
    return {"success": True, "remaining_count": len(remaining)}


@supplier_router.post("/bulk-delete")
async def bulk_delete_suppliers(req: BulkDeleteRequest, current_user: dict = Depends(auth.require_auth)):
    from main import _client_user, _now_utc
    user, client = _client_user(current_user)
    suppliers = client.setdefault("suppliers", [])
    remaining = [item for item in suppliers if item.get("id") not in set(req.supplier_ids)]
    deleted_count = len(suppliers) - len(remaining)
    client["suppliers"] = remaining
    client["supplier_count"] = len(remaining)
    client["updated_at"] = _now_utc()
    storage.write_audit(
        event_id=f"suppliers_{user['client_id']}",
        agent="SupplierManager",
        action="suppliers_bulk_deleted",
        input_summary=f"count={deleted_count}",
        output_summary=f"remaining={len(remaining)}",
        client_id=user["client_id"],
    )
    return {"success": True, "deleted_count": deleted_count, "remaining_count": len(remaining)}


@supplier_router.get("/export")
async def export_suppliers(current_user: dict = Depends(auth.require_auth)):
    from main import _client_user, _make_suppliers_workbook
    _, client = _client_user(current_user)
    wb = _make_suppliers_workbook(client)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    company_slug = client.get("company_name", "client").replace(" ", "_").lower()
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=disruptiq_suppliers_{company_slug}.xlsx"},
    )


@supplier_router.get("/health-scores")
async def supplier_health_scores(current_user: dict = Depends(auth.require_auth)):
    from main import _client_user
    _, client = _client_user(current_user)
    scores = []
    for supplier in client.get("suppliers", []):
        reliability = max(0, min(100, int(supplier.get("reliability", 85))))
        buffer_days = max(0, int(supplier.get("buffer_stock_days", 7)))
        sites = max(1, int(supplier.get("sites", 1)))
        proximity = max(1, min(10, int(supplier.get("proximity_score", 5))))
        score = min(100, round(reliability * 0.55 + min(buffer_days, 30) * 1.2 + min(sites, 5) * 6 + proximity * 2))
        trend = "stable"
        if score >= 80:
            trend = "improving"
        elif score < 50:
            trend = "declining"
        scores.append({
            "supplier_id": supplier.get("id"),
            "supplier_name": supplier.get("name"),
            "score": score,
            "trend": trend,
        })
    return {"scores": scores}


@supplier_router.get("/financial-health")
async def supplier_financial_health(current_user: dict = Depends(auth.require_auth)):
    """Section 2 Sprint: per-supplier financial health assessment.

    Computes an inferred 0-100 financial-health score per supplier from
    industry sector stress, distress-keyword news mentions (when NewsAPI is
    live), and operational proxies (reliability + buffer days).

    Empty state: returns a helpful message + empty suppliers/summary when
    the caller has uploaded no suppliers. Identical behaviour for demo
    clients (seed data) and real onboarded clients (uploaded data).
    """
    from main import _resolve_suppliers, _now_utc
    client_id = current_user["client_id"]
    suppliers = _resolve_suppliers(client_id)
    if not suppliers:
        return {
            "client_id": client_id,
            "suppliers": [],
            "summary": financial_signals.summarise_portfolio([]),
            "message": "Upload suppliers to see financial health analysis.",
            "last_updated_utc": _now_utc(),
        }

    records = []
    for supplier in suppliers:
        categories = supplier.get("categories") or ["Unknown"]
        primary_category = categories[0] if categories else "Unknown"
        health = financial_signals.compute_financial_health(
            supplier_name=supplier.get("name", ""),
            supplier_category=primary_category,
            reliability_pct=float(supplier.get("reliability", 80)),
            buffer_days=float(supplier.get("buffer_stock_days", 7)),
            news_alerts=[],
        )
        records.append({
            "supplier_id": supplier.get("id"),
            "supplier_name": supplier.get("name"),
            "zone": supplier.get("zone"),
            **health,
        })

    tier_order = {"Critical": 0, "At Risk": 1, "Watch": 2, "Stable": 3}
    records.sort(key=lambda r: tier_order.get(r.get("tier", "Stable"), 4))

    return {
        "client_id": client_id,
        "suppliers": records,
        "summary": financial_signals.summarise_portfolio(records),
        "last_updated_utc": _now_utc(),
    }


@supplier_router.get("/esg")
async def supplier_esg(current_user: dict = Depends(auth.require_auth)):
    """Section 4 Sprint: per-supplier ESG score and portfolio summary.

    Computes a 0-100 ESG composite from industry carbon intensity, zone
    climate risk, and industry labor risk. Returns supplier-level records
    + a portfolio summary (per-tier counts and average composite).
    """
    from main import _resolve_suppliers, clients_db, _now_utc
    client_id = current_user["client_id"]
    suppliers = _resolve_suppliers(client_id)
    if not suppliers:
        return {
            "client_id": client_id,
            "suppliers": [],
            "summary": esg_signals.summarise_esg_portfolio([]),
            "message": "Upload suppliers to see ESG analysis.",
            "last_updated_utc": _now_utc(),
        }

    client_industry = (clients_db.get(client_id) or {}).get("industry", "")

    records = []
    for supplier in suppliers:
        categories = supplier.get("categories") or []
        # Primary industry = supplier's first category if available;
        # otherwise the client's registered industry.
        industry = (categories[0] if categories else client_industry) or client_industry
        esg = esg_signals.compute_esg_score(
            supplier_name=supplier.get("name", ""),
            industry=industry,
            zone=supplier.get("zone", ""),
        )
        records.append({
            "supplier_id": supplier.get("id"),
            "supplier_name": supplier.get("name"),
            "zone": supplier.get("zone"),
            **esg,
        })

    tier_order = {"D": 0, "C": 1, "B": 2, "A": 3}
    records.sort(key=lambda r: tier_order.get(r.get("tier", "D"), 0))

    return {
        "client_id": client_id,
        "suppliers": records,
        "summary": esg_signals.summarise_esg_portfolio(records),
        "last_updated_utc": _now_utc(),
    }


@supplier_router.get("/compare")
async def compare_suppliers(ids: str, current_user: dict = Depends(auth.require_auth)):
    from main import _client_user
    _, client = _client_user(current_user)
    requested = {item.strip() for item in ids.split(",") if item.strip()}
    suppliers = [supplier for supplier in client.get("suppliers", []) if supplier.get("id") in requested]
    return {
        "suppliers": [
            {
                "id": supplier.get("id"),
                "name": supplier.get("name"),
                "zone": supplier.get("zone"),
                "categories": supplier.get("categories", []),
                "reliability": supplier.get("reliability", 85),
                "buffer_stock_days": supplier.get("buffer_stock_days", 7),
                "sites": supplier.get("sites", 1),
            }
            for supplier in suppliers
        ]
    }


@supplier_router.get("/trends")
async def supplier_trends(current_user: dict = Depends(auth.require_auth)):
    """Generate 30-day synthetic health score trends per supplier (Feature 2)."""
    from main import _resolve_suppliers, _now_utc
    client_id = current_user["client_id"]
    suppliers = _resolve_suppliers(client_id)
    result_suppliers = []

    for sup in suppliers:
        sup_id = sup.get("id", "")
        rng = random.Random(int(hashlib.md5(sup_id.encode()).hexdigest(), 16) % (2**32))

        reliability = max(0, min(100, int(sup.get("reliability", 85))))
        buffer_days = max(0, int(sup.get("buffer_stock_days", 7)))
        sites = max(1, int(sup.get("sites", 1)))
        proximity = max(1, min(10, int(sup.get("proximity_score", 5))))
        current_score = min(100, round(reliability * 0.55 + min(buffer_days, 30) * 1.2 + min(sites, 5) * 6 + proximity * 2))

        trend_data = []
        base = float(current_score)
        for i in range(29, -1, -1):
            date = (datetime.now(timezone.utc) - timedelta(days=i)).strftime("%Y-%m-%d")
            base += rng.gauss(0, 1.8)
            score = min(100, max(0, round(base, 1)))
            risk_tier = "Critical" if score >= 85 else "High" if score >= 70 else "Medium" if score >= 31 else "Low"
            trend_data.append({"date": date, "health_score": score, "risk_tier": risk_tier})

        result_suppliers.append({
            "id": sup_id,
            "name": sup.get("name", ""),
            "zone": sup.get("zone", ""),
            "current_health_score": current_score,
            "trend_data": trend_data,
        })

    return {
        "client_id": client_id,
        "suppliers": result_suppliers,
        "generated_at_utc": _now_utc(),
    }


@supplier_router.get("/anomalies")
async def supplier_anomalies(current_user: dict = Depends(auth.require_auth)):
    """Detect supplier anomalies: buffer, reliability, concentration risks (Feature 5c)."""
    from main import _resolve_suppliers, _now_utc
    client_id = current_user["client_id"]
    suppliers = _resolve_suppliers(client_id)
    result = agents.anomaly_detection_agent(suppliers)
    result["client_id"] = client_id
    result["generated_at_utc"] = _now_utc()
    return result
