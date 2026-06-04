"""
DisruptIQ V2 — Dataset generator.
Produces 10 industry-specific .xlsx supplier files covering both Indian and global zones.
Mix: some datasets < 30 suppliers (free-tier), some > 30 (premium).

Run:  python dataset/_generate.py
"""
import os, random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL   = PatternFill("solid", fgColor="1E1B4B")
HEADER_FONT   = Font(color="A78BFA", bold=True, size=11)
ROW_FILL_ALT  = PatternFill("solid", fgColor="0F1120")
ROW_FILL_EVEN = PatternFill("solid", fgColor="13152A")
CELL_FONT     = Font(color="CBD5E1", size=10)
BORDER_SIDE   = Side(style="thin", color="2D2F4A")

HEADERS = ["Supplier Name", "Zone", "Categories", "Buffer Stock Days",
           "Sites", "Reliability (%)", "Proximity Score (1-10)"]
COL_WIDTHS = [32, 18, 45, 20, 8, 18, 22]

INDIA_ZONES = [
    "Chennai","Mumbai","Delhi","Bengaluru","Pune","Kolkata",
    "Hyderabad","Ahmedabad","Coimbatore","Jaipur","Kochi",
    "Lucknow","Tamil Nadu","Gujarat","Rajasthan",
]
GLOBAL_ZONES = [
    "Shanghai","Shenzhen","Beijing","Tokyo","Seoul","Singapore",
    "Bangkok","Kuala Lumpur","Jakarta","Dubai","Frankfurt",
    "Rotterdam","London","Los Angeles","New York","Chicago",
    "Houston","São Paulo","Mexico City","Sydney","Melbourne",
    "Taipei","Stockholm","Mumbai",
]

DATASETS = [
    ("01_Automotive_Global_25_suppliers.xlsx",
     "Automotive — Global Supply Chain", GLOBAL_ZONES, 25,
     ["Chassis Parts","Engine Components","Wiring Harness","Brake Systems",
      "Transmission Parts","Steel Stampings","Forged Components","Sensors",
      "Electronic Control Units","Tyres","Plastics","Lighting Systems"]),

    ("02_Electronics_Global_35_suppliers.xlsx",
     "Consumer Electronics — Global", GLOBAL_ZONES, 35,
     ["Semiconductors","PCBs","Display Panels","Batteries","Connectors",
      "Memory Modules","Power ICs","Passive Components","Camera Modules",
      "Touchscreens","RF Modules","Audio Components","Thermal Management"]),

    ("03_Pharma_India_18_suppliers.xlsx",
     "Pharmaceuticals — India", INDIA_ZONES, 18,
     ["Active Pharmaceutical Ingredients","Excipients","Packaging",
      "Lab Reagents","Cold Chain Logistics","Clinical Supplies",
      "Bulk Drugs","Sterile Vials","Medical Devices","Enzymes"]),

    ("04_FMCG_Mixed_40_suppliers.xlsx",
     "FMCG — Global + India", GLOBAL_ZONES + INDIA_ZONES, 40,
     ["Packaging Materials","Fragrances","Surfactants","Palm Oil",
      "Sugar","Starch","Preservatives","Colorants","PET Resin",
      "Labels","Secondary Packaging","Cold Chain","Logistics"]),

    ("05_Aerospace_Global_22_suppliers.xlsx",
     "Aerospace & Defence — Global", GLOBAL_ZONES, 22,
     ["Titanium Alloys","Composite Panels","Avionics","Hydraulic Systems",
      "Landing Gear Components","Fasteners","Precision Machined Parts",
      "Fuel Systems","Electrical Wiring","Navigation Modules","Coatings"]),

    ("06_Renewable_Energy_Global_28_suppliers.xlsx",
     "Renewable Energy — Global", GLOBAL_ZONES, 28,
     ["Solar Cells","Wind Turbine Blades","Inverters","Battery Storage",
      "Structural Steel","Cables","Power Electronics","Control Systems",
      "Thermal Insulation","Gearboxes","Bearings","Foundation Steel"]),

    ("07_Food_Beverage_India_32_suppliers.xlsx",
     "Food & Beverage — India", INDIA_ZONES, 32,
     ["Raw Spices","Sugar Cane","Milk Solids","Wheat","Rice","Edible Oils",
      "Flavourings","Preservatives","Glass Bottles","Tetra Pak","Labels",
      "Cold Chain Logistics","Secondary Packaging","Starch","Emulsifiers"]),

    ("08_Chemicals_Mixed_20_suppliers.xlsx",
     "Specialty Chemicals — Mixed", GLOBAL_ZONES + INDIA_ZONES, 20,
     ["Solvents","Catalysts","Surfactants","Polymers","Adhesives",
      "Coatings","Industrial Gases","Acids","Base Chemicals",
      "Specialty Additives","Resins","Lubricants"]),

    ("09_Logistics_3PL_Global_15_suppliers.xlsx",
     "3PL / Logistics — Global", GLOBAL_ZONES, 15,
     ["Ocean Freight","Air Freight","Last-Mile Delivery","Cold Chain",
      "Customs Brokerage","Warehouse Management","Cross-Docking",
      "Express Courier","Port Handling","Rail Freight","Packaging"]),

    ("10_Medical_Devices_Global_38_suppliers.xlsx",
     "Medical Devices — Global", GLOBAL_ZONES, 38,
     ["Surgical Instruments","Imaging Components","Implant Materials",
      "Electronic Sensors","Sterilisation Supplies","Disposables",
      "Diagnostic Reagents","Plastics","Optical Lenses","Batteries",
      "PCBs","RF Components","Coatings","Packaging Materials"]),
]

PREFIXES = [
    "Apex","Nordic","Pacific","Eurotech","Global","Atlas","Meridian","Summit",
    "Pinnacle","Nexus","Titan","Vanguard","Horizon","Pioneer","Quantum",
    "Delta","Orion","Zenith","Eclipse","Fusion","Helix","Nova","Stratos",
    "Omni","Prism","Vertex","Solaris","Cascade","Alliance","Vector",
]
SUFFIXES = [
    "Industries","Manufacturing","Components","Systems","Solutions",
    "Technologies","Enterprises","Supplies","Group","Corp","International",
    "Holdings","Partners","Dynamics","Precision","Works","Labs","Co",
]

def make_name(used):
    for _ in range(200):
        n = f"{random.choice(PREFIXES)} {random.choice(SUFFIXES)}"
        if n not in used:
            used.add(n)
            return n
    return f"Supplier {len(used)+1}"

def rand_cats(pool, n=2):
    return ", ".join(random.sample(pool, min(n, len(pool))))

def write_dataset(path, label, zones, count, cats_pool):
    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    # Row 1 = column headers (no title row)
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            bottom=Side(style="medium", color="7C6BFF"),
            right=BORDER_SIDE,
        )
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    used_names = set()
    for i in range(count):
        row = i + 2
        zone  = random.choice(zones)
        fill  = ROW_FILL_ALT if i % 2 == 0 else ROW_FILL_EVEN
        data  = [
            make_name(used_names),
            zone,
            rand_cats(cats_pool, random.randint(1, 3)),
            random.randint(3, 45),
            random.randint(1, 5),
            random.randint(65, 99),
            random.randint(1, 10),
        ]
        for col, val in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = CELL_FONT
            cell.fill = fill
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if col > 2 else "left"
            )
            cell.border = Border(right=BORDER_SIDE, bottom=BORDER_SIDE)
        ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"  {count:>3} suppliers  ->  {os.path.basename(path)}")


if __name__ == "__main__":
    out_dir = os.path.dirname(os.path.abspath(__file__))
    print(f"Writing to {out_dir}\n")
    for fname, label, zones, count, cats in DATASETS:
        write_dataset(os.path.join(out_dir, fname), label, zones, count, cats)
    print(f"\nDone — {len(DATASETS)} files.")
