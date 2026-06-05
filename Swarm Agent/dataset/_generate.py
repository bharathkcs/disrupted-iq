"""Generate 10 industry-specific supplier Excel files for client onboarding tests.

Run from inside this folder:
    cd "e:\\Swarm\\Swarm Agent\\dataset"
    python _generate.py

Output: 10 .xlsx files, each matching the upload schema:
  Supplier Name* | Zone* | Categories* (comma-separated) | Buffer Stock Days | Sites | Reliability (%) | Proximity Score (1-10)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADERS = [
    "Supplier Name*",
    "Zone*",
    "Categories* (comma-separated)",
    "Buffer Stock Days",
    "Sites",
    "Reliability (%)",
    "Proximity Score (1-10)",
]

HEADER_FILL = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ROW_FILL = PatternFill(start_color="0D1123", end_color="0D1123", fill_type="solid")
ROW_FONT = Font(color="E2E8F0", size=10)
ALIGN_CENTER = Alignment(vertical="center", horizontal="center")
ALIGN_LEFT = Alignment(vertical="center", horizontal="left")


def write_workbook(filename: str, rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    # Header row
    for col_idx, header in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.fill = ROW_FILL
            c.font = ROW_FONT
            c.alignment = ALIGN_LEFT if col_idx <= 3 else ALIGN_CENTER

    # Column widths
    widths = [28, 16, 36, 18, 8, 16, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(filename)
    print(f"  [OK] {filename} ({len(rows)} suppliers)")


# ════════════════════════════════════════════════════════════════════════════
# 1. AUTOMOTIVE — 15 suppliers (Tier-1, Tier-2, raw materials)
# ════════════════════════════════════════════════════════════════════════════
automotive = [
    ["Bosch India",                "Pune",        "Brake Systems, Fuel Injection, ECU",        21, 4, 96,  9],
    ["Bharat Forge",               "Pune",        "Forged Components, Crankshafts",            18, 3, 94,  9],
    ["Sundaram Fasteners",         "Chennai",     "Fasteners, Powder Metallurgy",              14, 2, 92,  8],
    ["Motherson Sumi",             "Noida",       "Wiring Harness, Plastic Components",        12, 5, 91,  6],
    ["Mahle India",                "Bengaluru",   "Pistons, Engine Components",                15, 2, 93,  7],
    ["JBM Auto",                   "Delhi",       "Sheet Metal, Body Panels",                  10, 3, 88,  7],
    ["Endurance Technologies",     "Aurangabad",  "Suspension, Brakes, Clutch",                16, 4, 92,  6],
    ["Apollo Tyres",               "Kochi",       "Tyres, Rubber Components",                  22, 3, 90,  5],
    ["Tata Steel",                 "Mumbai",      "Steel Coil, Galvanized Sheet",              28, 6, 95,  8],
    ["Hindalco Industries",        "Mumbai",      "Aluminum Sheets, Castings",                 25, 4, 93,  8],
    ["Continental India",          "Bengaluru",   "Tyres, Electronics, Brake Systems",         18, 2, 94,  7],
    ["Bharat Petroleum",           "Mumbai",      "Lubricants, Engine Oil",                    20, 5, 89,  8],
    ["SKF India",                  "Pune",        "Bearings, Sealing Solutions",               16, 2, 95,  9],
    ["Denso Kirloskar",            "Bengaluru",   "AC Compressors, Spark Plugs, Sensors",      14, 2, 93,  7],
    ["JK Tyre & Industries",       "Delhi",       "Tyres, Tubes",                              20, 4, 88,  6],
]

# ════════════════════════════════════════════════════════════════════════════
# 2. ELECTRONICS — 20 suppliers (semiconductors, PCB, displays)
# ════════════════════════════════════════════════════════════════════════════
electronics = [
    ["TSMC",                       "Taiwan",      "Semiconductors, 5nm Wafers",                45, 3, 98,  1],
    ["Samsung Electronics",        "South Korea", "DRAM, NAND Flash, Displays",                30, 8, 96,  1],
    ["SK Hynix",                   "South Korea", "DRAM, NAND, Memory ICs",                    35, 4, 95,  1],
    ["Foxconn India",              "Chennai",     "PCB Assembly, EMS, Casings",                14, 5, 92,  8],
    ["Wistron",                    "Bengaluru",   "PCB Assembly, Mobile Components",           12, 3, 90,  8],
    ["Dixon Technologies",         "Noida",       "EMS, Set-top Boxes, Mobiles",               10, 4, 88,  7],
    ["Amber Enterprises",          "Delhi",       "Air Conditioning Electronics, IDU/ODU",     15, 3, 91,  7],
    ["Vu Technologies",             "Mumbai",     "LED TVs, Displays",                          8, 2, 86,  8],
    ["Syrma SGS",                  "Chennai",     "PCB Assembly, IoT Modules",                 14, 4, 89,  8],
    ["Kaynes Technology",          "Mysuru",      "Box-Build Electronics, PCBA",               12, 3, 90,  7],
    ["Bharat Electronics (BEL)",   "Bengaluru",   "Defense Electronics, Radars, Comms",        45, 6, 96,  8],
    ["TE Connectivity",            "Bengaluru",   "Connectors, Sensors, Cables",               18, 2, 93,  7],
    ["LG Display",                 "South Korea", "OLED Panels, LCD Modules",                  28, 3, 94,  1],
    ["Murata Manufacturing",       "Japan",       "Capacitors, Wireless Modules",              35, 4, 97,  1],
    ["Texas Instruments",          "USA",         "Power Management IC, MCU",                  50, 5, 96,  1],
    ["Infineon Technologies",      "Germany",     "Power Semiconductors, MCU",                 45, 4, 95,  1],
    ["STMicroelectronics",         "France",      "MCU, Power Devices, Sensors",               40, 3, 94,  1],
    ["Tata Elxsi",                 "Bengaluru",   "Embedded Software, Design Services",        21, 2, 92,  8],
    ["Centum Electronics",         "Bengaluru",   "PCBA, Cable Harness, Aerospace",            16, 2, 89,  7],
    ["Optiemus Electronics",       "Noida",       "Mobile PCBA, Wearables",                    10, 2, 87,  7],
]

# ════════════════════════════════════════════════════════════════════════════
# 3. PHARMACEUTICAL — 10 suppliers (APIs, formulations, packaging)
# ════════════════════════════════════════════════════════════════════════════
pharmaceutical = [
    ["Divi's Laboratories",        "Hyderabad",   "API, Custom Synthesis, Generics",           30, 3, 96,  7],
    ["Dr. Reddy's Laboratories",   "Hyderabad",   "API, Formulations, Generics",               28, 5, 95,  7],
    ["Aurobindo Pharma",           "Hyderabad",   "API, Antibiotics, Antiretrovirals",         25, 4, 94,  7],
    ["Lupin Limited",              "Mumbai",      "API, Respiratory Formulations",             24, 4, 93,  8],
    ["Sun Pharma",                 "Mumbai",      "API, Specialty Generics",                   30, 6, 96,  8],
    ["Granules India",             "Hyderabad",   "API, Paracetamol, Metformin",               22, 3, 92,  7],
    ["Laurus Labs",                "Hyderabad",   "API, ARV, Oncology",                        20, 3, 91,  7],
    ["Cadila Pharmaceuticals",     "Ahmedabad",   "API, Formulations, Vaccines",               25, 4, 93,  8],
    ["Schott Kaisha",              "Mumbai",      "Pharma Glass Packaging, Vials, Ampoules",   28, 2, 90,  8],
    ["Bilcare Limited",            "Pune",        "Pharma Packaging Films, Foils",             18, 2, 88,  7],
]

# ════════════════════════════════════════════════════════════════════════════
# 4. FMCG — 25 suppliers (ingredients, packaging, distribution)
# ════════════════════════════════════════════════════════════════════════════
fmcg = [
    ["Britannia Industries",       "Bengaluru",   "Wheat Flour, Sugar, Edible Oil",            14, 5, 92,  7],
    ["Adani Wilmar",               "Ahmedabad",   "Edible Oil, Wheat, Rice",                   21, 6, 94,  8],
    ["Patanjali Foods",            "Haridwar",    "Edible Oil, Soya, Cosmetics Ingredients",   18, 4, 88,  6],
    ["EID Parry",                  "Chennai",     "Sugar, Bio-products, Ethanol",              28, 3, 92,  8],
    ["Tata Consumer Products",     "Mumbai",      "Tea, Salt, Pulses",                         21, 5, 93,  8],
    ["Marico",                     "Mumbai",      "Coconut Oil, Edible Oil, Hair Care",        16, 4, 91,  8],
    ["Dabur India",                "Ghaziabad",   "Honey, Herbal Ingredients, Juices",         18, 3, 92,  6],
    ["Godrej Consumer Products",   "Mumbai",      "Soap Noodles, Fragrances, Packaging",       14, 4, 90,  8],
    ["Emami Limited",              "Kolkata",     "Herbal Ingredients, Hair Oil Base",         16, 2, 89,  6],
    ["VST Industries",             "Hyderabad",   "Tobacco Leaf, Packaging Materials",         24, 2, 88,  7],
    ["Ruchi Soya (Patanjali)",     "Indore",      "Edible Oil, Soya, Pulses",                  20, 3, 87,  6],
    ["Sona BLW",                   "Gurgaon",     "Packaging Material, Films",                 12, 3, 89,  6],
    ["Uflex",                      "Noida",       "Flexible Packaging, Films, Foils",          14, 4, 91,  7],
    ["Essel Propack",              "Mumbai",      "Laminated Tubes, Specialty Packaging",      14, 3, 90,  8],
    ["Manjushree Technopack",      "Bengaluru",   "Plastic Containers, Bottles",               12, 4, 88,  7],
    ["Borosil Renewables",         "Gujarat",      "Glass Packaging, Bottles",                  16, 2, 89,  7],
    ["AGI Glaspac",                "Hyderabad",   "Glass Containers, Beverage Bottles",        18, 2, 90,  7],
    ["TCPL Packaging",             "Mumbai",      "Folding Cartons, Printed Packaging",        14, 3, 88,  8],
    ["ITC Limited",                "Kolkata",     "Tobacco, Paper, Hotel & Foods",             21, 8, 94,  6],
    ["Nestle India",               "Gurgaon",     "Coffee, Dairy, Chocolate Ingredients",      18, 4, 95,  6],
    ["Hindustan Unilever",         "Mumbai",      "Vegetable Oil, Surfactants, Fragrances",    21, 7, 96,  8],
    ["Procter & Gamble",           "Mumbai",      "Surfactants, Polymers, Fragrances",         24, 5, 95,  8],
    ["Colgate-Palmolive",          "Mumbai",      "Toothpaste Base, Soap Noodles",             18, 3, 93,  8],
    ["Heinz India",                "Mumbai",      "Tomato Paste, Vinegar, Spices",             16, 2, 91,  8],
    ["Mother Dairy",               "Delhi",       "Dairy Ingredients, Milk Powder",            10, 4, 90,  6],
]

# ════════════════════════════════════════════════════════════════════════════
# 5. LOGISTICS — 30 suppliers (transport, warehousing, freight, customs)
# ════════════════════════════════════════════════════════════════════════════
logistics = [
    ["Blue Dart Express",          "Mumbai",      "Air Express, Same-Day, COD",                 7, 8, 96,  8],
    ["DTDC Express",               "Bengaluru",   "Surface Transport, Ecommerce Logistics",     5, 6, 88,  8],
    ["Delhivery",                  "Gurgaon",     "Ecommerce Logistics, B2B Freight",           4, 9, 90,  6],
    ["TCI Express",                "Gurgaon",     "Surface Express, Less-than-Truckload",       6, 7, 91,  6],
    ["VRL Logistics",              "Hubli",       "Trucking, LTL, Parcel",                      8, 5, 89,  5],
    ["Allcargo Logistics",         "Mumbai",      "Multimodal, NVOCC, CFS",                    14, 4, 93,  8],
    ["Container Corporation",      "Delhi",       "Rail Logistics, ICD, Containers",           21, 8, 94,  6],
    ["GATI Limited",               "Hyderabad",   "Express Cargo, Domestic Freight",            7, 6, 87,  7],
    ["Mahindra Logistics",         "Mumbai",      "3PL, Warehousing, Last-Mile",               10, 5, 90,  8],
    ["Future Supply Chain",        "Mumbai",      "Cold Chain, Retail Logistics",              12, 4, 89,  8],
    ["Maersk India",               "Mumbai",      "Ocean Freight, Container Shipping",         30, 3, 95,  8],
    ["MSC India",                  "Mumbai",      "Container Shipping, Reefer",                30, 3, 94,  8],
    ["DHL Express India",          "Mumbai",      "International Air Express, Customs",        14, 4, 96,  8],
    ["FedEx India",                "Mumbai",      "Air Express, Time-Definite, Customs",       14, 4, 95,  8],
    ["UPS India",                  "Mumbai",      "Air Express, International Brokerage",      14, 3, 95,  8],
    ["Bharat Petroleum",           "Mumbai",      "Fuel, Diesel Distribution",                  3, 5, 92,  8],
    ["Indian Oil Corp",            "Delhi",       "Diesel, Bulk Fuel, Lubricants",              3, 8, 93,  6],
    ["Reliance Petroleum",         "Mumbai",      "Diesel, Bunker Fuel",                        3, 6, 91,  8],
    ["JNPT (Jawaharlal Nehru)",    "Mumbai",      "Port Operations, Container Handling",       30, 1, 92,  9],
    ["Adani Ports & SEZ",          "Mumbai",      "Port Operations, Terminal Handling",        30, 7, 94,  8],
    ["Chennai Port Trust",         "Chennai",     "Port Operations, Bulk Cargo",               25, 1, 90,  9],
    ["Air India SATS",             "Delhi",       "Cargo Handling, Ground Services",            7, 4, 90,  6],
    ["Concorde Air Logistics",     "Bengaluru",   "Air Freight, Cargo Handling",                7, 2, 87,  8],
    ["JM Baxi & Co",               "Mumbai",      "Customs Clearance, NVOCC",                  10, 3, 92,  8],
    ["Aegis Logistics",            "Mumbai",      "LPG Logistics, Liquid Terminals",           21, 4, 91,  8],
    ["Snowman Logistics",          "Bengaluru",   "Cold Chain Warehousing, Reefer Transport",  10, 5, 88,  8],
    ["Coldex Logistics",           "Delhi",       "Cold Chain, Refrigerated Transport",        12, 3, 87,  6],
    ["TVS Supply Chain",           "Chennai",     "Automotive Logistics, In-plant",            10, 4, 90,  8],
    ["Safexpress",                 "Delhi",       "Surface Express, B2B Logistics",             8, 6, 89,  6],
    ["Spoton Logistics",           "Bengaluru",   "B2B Express, LTL",                           7, 5, 88,  7],
]

# ════════════════════════════════════════════════════════════════════════════
# 6. TEXTILE / RUBBER / FIBER — 12 suppliers (yarn, dyes, chemicals)
# ════════════════════════════════════════════════════════════════════════════
textile = [
    ["Welspun India",              "Vapi",        "Home Textiles, Towels, Sheets",             21, 4, 93,  7],
    ["Trident Limited",            "Ludhiana",    "Yarn, Bed Linen, Paper",                    18, 3, 91,  6],
    ["Arvind Limited",             "Ahmedabad",   "Denim Fabric, Dyes",                        18, 4, 92,  8],
    ["Vardhman Textiles",          "Ludhiana",    "Yarn, Sewing Thread, Fabric",               21, 3, 90,  6],
    ["Raymond Limited",            "Mumbai",      "Worsted Fabric, Suiting",                   24, 4, 92,  8],
    ["Bombay Dyeing",              "Mumbai",      "Cotton Yarn, Fabric, Polyester Staple",     18, 2, 88,  8],
    ["Shahi Exports",              "Bengaluru",   "Garments, Knitted Fabric",                  14, 5, 90,  8],
    ["Teijin Frontier",            "Thailand",    "Yarn Fibre, Polyester Filament",            45, 1, 95,  1],
    ["Kuraray",                    "Japan",       "PVA Fibre, Specialty Polymers",             60, 1, 95,  1],
    ["Kumho Petrochemical",        "South Korea", "Synthetic Rubber, Latex",                   60, 1, 95,  1],
    ["Indian Dyestuff Industries", "Mumbai",      "Dyes, Pigments, Auxiliaries",               21, 2, 87,  8],
    ["Nocil",                      "Mumbai",      "Rubber Chemicals, Accelerators",            12, 1, 85,  8],
]

# ════════════════════════════════════════════════════════════════════════════
# 7. STEEL & HEAVY MANUFACTURING — 18 suppliers
# ════════════════════════════════════════════════════════════════════════════
steel = [
    ["Tata Steel",                 "Jamshedpur",  "Steel Coil, Galvanized, TMT Bars",          30, 6, 96,  6],
    ["JSW Steel",                  "Mumbai",      "Hot Rolled Coil, Steel Plates",             28, 5, 95,  8],
    ["SAIL (Steel Authority)",     "Delhi",       "Steel Plates, Beams, Rails",                35, 8, 92,  6],
    ["Jindal Steel & Power",       "Delhi",       "Steel, Power, Mining",                      32, 5, 91,  6],
    ["Vedanta Limited",            "Mumbai",      "Aluminum, Zinc, Copper, Iron Ore",          28, 7, 93,  8],
    ["Hindalco",                   "Mumbai",      "Aluminum Sheets, Extrusions, Castings",     25, 5, 94,  8],
    ["NMDC Limited",               "Hyderabad",   "Iron Ore, Sponge Iron, Pellets",            30, 4, 90,  7],
    ["Adani Enterprises",          "Ahmedabad",   "Coal, Iron Ore, Mining",                    28, 6, 89,  8],
    ["Coal India Limited",         "Kolkata",     "Coal, Coking Coal",                         21, 9, 88,  6],
    ["L&T Heavy Engineering",      "Mumbai",      "Pressure Vessels, Heat Exchangers, Forging", 35, 4, 95,  8],
    ["Bharat Heavy Electricals",   "Delhi",       "Power Equipment, Boilers, Turbines",        45, 6, 93,  6],
    ["Welspun Corp",               "Mumbai",      "Steel Pipes, Plates",                       25, 3, 91,  8],
    ["Maharashtra Seamless",       "Mumbai",      "Seamless Pipes, Tubes",                     21, 2, 89,  8],
    ["APL Apollo Tubes",           "Delhi",       "Steel Tubes, Hollow Sections",              18, 4, 90,  6],
    ["Sandvik Asia",               "Pune",        "Specialty Steel, Cutting Tools",            28, 2, 94,  9],
    ["Kalyani Steels",             "Pune",        "Forging Steel, Alloy Steel",                21, 2, 91,  9],
    ["Mukand Limited",             "Mumbai",      "Stainless Steel, Alloy Steel",              18, 2, 88,  8],
    ["RINL Vizag",                 "Visakhapatnam","Steel Long Products, TMT",                 24, 3, 89,  6],
]

# ════════════════════════════════════════════════════════════════════════════
# 8. AEROSPACE & DEFENSE — 8 suppliers (precision, critical)
# ════════════════════════════════════════════════════════════════════════════
aerospace = [
    ["Hindustan Aeronautics (HAL)","Bengaluru",   "Aircraft, Helicopters, Engines",            90, 5, 96,  9],
    ["Bharat Electronics",         "Bengaluru",   "Radar, Avionics, Defense Electronics",      60, 4, 95,  9],
    ["Tata Advanced Systems",      "Hyderabad",   "Aerostructures, Composites, UAVs",          60, 3, 94,  7],
    ["Mahindra Defence Systems",   "Pune",        "Armored Vehicles, Naval Systems",           45, 2, 92,  9],
    ["L&T Defence",                "Mumbai",      "Submarine Systems, Missile Launchers",      60, 3, 95,  8],
    ["Dynamatic Technologies",     "Bengaluru",   "Aerospace Precision Components",            45, 2, 93,  9],
    ["GE Aviation India",          "Bengaluru",   "Jet Engine Components, R&D",                90, 3, 96,  9],
    ["Honeywell Aerospace",        "Bengaluru",   "Avionics, Cockpit Systems",                 75, 2, 95,  9],
]

# ════════════════════════════════════════════════════════════════════════════
# 9. FOOD & BEVERAGE — 22 suppliers (dairy, grains, beverages, ingredients)
# ════════════════════════════════════════════════════════════════════════════
food = [
    ["Amul (GCMMF)",               "Anand",       "Milk Powder, Butter, Cheese",                7, 8, 94,  6],
    ["Mother Dairy",               "Delhi",       "Liquid Milk, Curd, Ice Cream",               5, 5, 92,  6],
    ["Heritage Foods",             "Hyderabad",   "Dairy, Curd, Milk Powder",                   7, 4, 90,  7],
    ["Hatsun Agro",                "Chennai",     "Dairy, Ice Cream, Curd",                     7, 6, 91,  8],
    ["Britannia Industries",       "Bengaluru",   "Biscuits, Bread, Dairy",                    14, 5, 93,  7],
    ["Parle Products",             "Mumbai",      "Biscuits, Confectionery",                   14, 4, 92,  8],
    ["ITC Foods",                  "Kolkata",     "Atta, Snacks, Spices",                      18, 5, 94,  6],
    ["Bikanervala Foods",          "Delhi",       "Sweets, Snacks, Frozen Foods",              10, 3, 89,  6],
    ["Haldiram's",                 "Delhi",       "Snacks, Sweets, Namkeen",                   12, 4, 91,  6],
    ["Nestle India",               "Gurgaon",     "Coffee, Maggi, Chocolate, Dairy",           21, 5, 96,  6],
    ["Coca-Cola India",            "Gurgaon",     "Concentrates, Sugar, PET",                  14, 6, 95,  6],
    ["PepsiCo India",              "Gurgaon",     "Concentrates, Snacks, PET",                 14, 5, 94,  6],
    ["Bisleri International",      "Mumbai",      "Mineral Water, PET Bottles",                10, 5, 90,  8],
    ["Tata Coffee",                "Bengaluru",   "Coffee Beans, Instant Coffee",              30, 3, 92,  7],
    ["McCormick India",            "Hyderabad",   "Spices, Seasonings, Flavorings",            21, 2, 91,  7],
    ["Olam International",         "Mumbai",      "Cocoa, Coffee, Cashew, Rice",               25, 4, 93,  8],
    ["Cargill India",              "Gurgaon",     "Edible Oil, Cocoa, Animal Feed",            21, 5, 94,  6],
    ["ADM India",                  "Mumbai",      "Soybean, Cocoa, Sweeteners",                21, 3, 92,  8],
    ["Britannia Bel Foods",        "Bengaluru",   "Cheese, Dairy Products",                    14, 2, 89,  7],
    ["Allana Group",               "Mumbai",      "Meat, Frozen Food, Marine",                 21, 4, 91,  8],
    ["Suguna Foods",               "Coimbatore",  "Poultry, Eggs, Animal Feed",                14, 5, 90,  7],
    ["Godrej Tyson Foods",         "Mumbai",      "Poultry, Frozen Protein",                   10, 3, 88,  8],
]

# ════════════════════════════════════════════════════════════════════════════
# 10. CHEMICALS — 16 suppliers (specialty, bulk, petrochemicals)
# ════════════════════════════════════════════════════════════════════════════
chemicals = [
    ["Reliance Industries",        "Mumbai",      "Polymers, Petrochemicals, PVC",             28, 6, 96,  8],
    ["GAIL India",                 "Delhi",       "Natural Gas, Petrochemicals",               21, 5, 94,  6],
    ["ONGC Petro Additions",       "Vadodara",    "Polyethylene, Polypropylene",               25, 2, 92,  8],
    ["Tata Chemicals",             "Mumbai",      "Soda Ash, Salt, Specialty Chemicals",       21, 5, 93,  8],
    ["UPL Limited",                "Mumbai",      "Agrochemicals, Crop Protection",            28, 6, 92,  8],
    ["Pidilite Industries",        "Mumbai",      "Adhesives, Resins, Sealants",               21, 4, 94,  8],
    ["SRF Limited",                "Gurgaon",     "Specialty Chemicals, Fluorochemicals",      24, 4, 93,  6],
    ["Aarti Industries",           "Mumbai",      "Specialty Chemicals, Pharma Intermediates", 21, 5, 92,  8],
    ["Vinati Organics",            "Mumbai",      "ATBS, IBB, Specialty Chemicals",            24, 2, 91,  8],
    ["Galaxy Surfactants",         "Mumbai",      "Surfactants, Personal Care Chemicals",      18, 3, 90,  8],
    ["BASF India",                 "Mumbai",      "Specialty Chemicals, Catalysts, Pigments",  30, 3, 95,  8],
    ["Dow Chemical India",         "Mumbai",      "Polymers, Coatings, Polyurethanes",         35, 2, 95,  8],
    ["Lanxess India",              "Mumbai",      "Rubber Chemicals, Engineering Plastics",    28, 2, 93,  8],
    ["Solvay India",               "Mumbai",      "Specialty Polymers, Surfactants",           30, 2, 94,  8],
    ["Deepak Nitrite",             "Vadodara",    "Sodium Nitrite, Phenolics, Specialty",      21, 3, 91,  8],
    ["Atul Limited",               "Valsad",      "Dyes, Pesticides, Specialty Chemicals",     24, 2, 90,  8],
]


DATASETS = [
    ("01_Automotive_15_suppliers.xlsx",          automotive),
    ("02_Electronics_20_suppliers.xlsx",         electronics),
    ("03_Pharmaceutical_10_suppliers.xlsx",      pharmaceutical),
    ("04_FMCG_25_suppliers.xlsx",                fmcg),
    ("05_Logistics_30_suppliers.xlsx",           logistics),
    ("06_Textile_Rubber_12_suppliers.xlsx",      textile),
    ("07_Steel_Manufacturing_18_suppliers.xlsx", steel),
    ("08_Aerospace_8_suppliers.xlsx",            aerospace),
    ("09_FoodBeverage_22_suppliers.xlsx",        food),
    ("10_Chemicals_16_suppliers.xlsx",           chemicals),
]


if __name__ == "__main__":
    print("Generating 10 industry-specific supplier datasets...\n")
    for filename, rows in DATASETS:
        write_workbook(filename, rows)
    print(f"\n  All {len(DATASETS)} datasets generated successfully.")
